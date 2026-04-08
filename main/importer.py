"""
XLSX mailing importer.

This module is intentionally decoupled from Django management commands so it
can be used in tests, scripts, or future API endpoints without modification.

Public interface
────────────────
    result = import_mailings_from_xlsx(path, send=True)
    print(result)   # ImportResult with counts

Column contract (first row must be a header row)
─────────────────────────────────────────────────
    external_id | user_id | email | subject | message
"""

from __future__ import annotations

import logging
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterator

import openpyxl
from django.core.exceptions import ValidationError
from django.core.validators import validate_email

from .models import MailingRecord

logger = logging.getLogger(__name__)

# Expected column names (order-independent; matched by header value).
REQUIRED_COLUMNS = frozenset({"external_id", "user_id", "email", "subject", "message"})


# ---------------------------------------------------------------------------
# Result container
# ---------------------------------------------------------------------------


@dataclass
class ImportResult:
    """Aggregated outcome of a single import run."""

    total_rows: int = 0
    created:    int = 0
    skipped:    int = 0   # rows whose external_id already exists in the DB
    errors:     int = 0
    error_details: list[str] = field(default_factory=list)

    def __str__(self) -> str:
        lines = [
            "── Mailing Import Result ──────────────────",
            f"  Processed rows : {self.total_rows}",
            f"  Created        : {self.created}",
            f"  Skipped        : {self.skipped}",
            f"  Errors         : {self.errors}",
        ]
        if self.error_details:
            lines.append("  Error details:")
            for detail in self.error_details:
                lines.append(f"    • {detail}")
        lines.append("───────────────────────────────────────────")
        return "\n".join(lines)


# ---------------------------------------------------------------------------
# Row-level helpers
# ---------------------------------------------------------------------------


@dataclass
class _Row:
    """Parsed and validated data for a single spreadsheet row."""

    external_id: str
    user_id: int
    email: str
    subject: str
    message: str


def _parse_row(row_dict: dict[str, str | None], row_num: int) -> _Row:
    """
    Parse and validate a single row dictionary.

    Raises ``ValueError`` with a human-readable message on any validation
    failure so the importer can collect errors without aborting.
    """
    errors: list[str] = []

    external_id = str(row_dict.get("external_id") or "").strip()
    if not external_id:
        errors.append("external_id is empty")

    raw_user_id = row_dict.get("user_id")
    user_id: int | None = None
    try:
        user_id = int(raw_user_id)  # type: ignore[arg-type]
        if user_id <= 0:
            raise ValueError
    except (TypeError, ValueError):
        errors.append(f"user_id must be a positive integer, got {raw_user_id!r}")

    email = str(row_dict.get("email") or "").strip()
    try:
        validate_email(email)
    except ValidationError:
        errors.append(f"invalid email address {email!r}")

    subject = str(row_dict.get("subject") or "").strip()
    if not subject:
        errors.append("subject is empty")

    message = str(row_dict.get("message") or "").strip()
    if not message:
        errors.append("message is empty")

    if errors:
        raise ValueError(f"Row {row_num}: " + "; ".join(errors))

    return _Row(
        external_id=external_id,
        user_id=user_id,       # type: ignore[arg-type]
        email=email,
        subject=subject,
        message=message,
    )


# ---------------------------------------------------------------------------
# XLSX reader
# ---------------------------------------------------------------------------


def _iter_rows(path: Path) -> Iterator[tuple[int, dict[str, str | None]]]:
    """
    Yield ``(row_number, row_dict)`` for every data row in the workbook.

    Row numbers are 1-based and account for the header row, so the first data
    row is reported as row 2 — matching the visual row number in Excel.

    Raises ``ValueError`` if the file is unreadable or has missing columns.
    """
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"Cannot open workbook {path}: {exc}") from exc

    ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)
    try:
        raw_header = next(rows_iter)
    except StopIteration:
        raise ValueError("The workbook is empty — no header row found.")

    headers = [str(h).strip().lower() if h is not None else "" for h in raw_header]

    missing = REQUIRED_COLUMNS - set(headers)
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(sorted(missing))}")

    for row_num, raw_values in enumerate(rows_iter, start=2):
        row_dict = dict(zip(headers, (str(v) if v is not None else None for v in raw_values)))
        yield row_num, row_dict

    wb.close()


# ---------------------------------------------------------------------------
# Public import function
# ---------------------------------------------------------------------------


def import_mailings_from_xlsx(path: Path | str, *, send: bool = True) -> ImportResult:
    """
    Import mailing records from an XLSX file and optionally enqueue delivery.

    Parameters
    ----------
    path:
        Filesystem path to the ``.xlsx`` file.
    send:
        When ``True`` (default), a Celery task is dispatched for every newly
        created record.  Pass ``False`` in tests or dry-run scenarios.

    Returns
    -------
    ImportResult
        Aggregated counts for the calling command to display.
    """
    path = Path(path)
    result = ImportResult()

    # Pre-load existing external_ids into a set for O(1) duplicate checks
    # instead of one SELECT per row.
    existing_ids: set[str] = set(
        MailingRecord.objects.values_list("external_id", flat=True)
    )

    records_to_create: list[MailingRecord] = []
    pending_rows: list[_Row]              = []

    for row_num, row_dict in _iter_rows(path):
        result.total_rows += 1

        try:
            row = _parse_row(row_dict, row_num)
        except ValueError as exc:
            result.errors += 1
            result.error_details.append(str(exc))
            logger.warning("Import validation error: %s", exc)
            continue

        if row.external_id in existing_ids:
            result.skipped += 1
            logger.debug("Skipping duplicate external_id=%s", row.external_id)
            continue

        # Guard against duplicates within the same file.
        existing_ids.add(row.external_id)

        records_to_create.append(
            MailingRecord(
                external_id=row.external_id,
                user_id=row.user_id,
                email=row.email,
                subject=row.subject,
                message=row.message,
            )
        )
        pending_rows.append(row)

    # Bulk-insert all new records in a single query.
    created_records = MailingRecord.objects.bulk_create(records_to_create)
    result.created = len(created_records)

    # Dispatch Celery tasks for newly created records.
    if send and created_records:
        from .tasks import dispatch_mailing  # local import to avoid circular deps

        for record in created_records:
            dispatch_mailing.delay(record.pk)
            logger.debug("Enqueued dispatch_mailing for record pk=%d", record.pk)

    return result
